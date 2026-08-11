#!/usr/bin/env python3
"""
Convert a simple passenger list (No., Full Name, Date of Birth, Passport No.,
Expiry, Room) into the "Pax Data" workbook format used for pax uploads.

Usage:
    python convert.py INPUT.xlsx [-o OUTPUT.xlsx] [--year YYYY] [--date DD.MM.YYYY]

The travel date used to compute passenger age (Adult/Child/Infant) is taken,
in order of priority:
    1. --date, if given
    2. DD.MM found in the input filename, combined with --year
       (--year defaults to the current year)
"""
import argparse
import datetime as dt
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PAX_DATA_HEADERS = [
    "Sequence",
    "Traveling With",
    "Pax Type",
    "Title",
    "First Name",
    "Last Name",
    "Gender",
    "DOB (dd/mm/yyyy)",
    "Nationality",
    "Mobile Number",
    "Email",
    "ID Number",
    "Passport Number",
    "Passport Expiry (dd/mm/yyyy)",
    "Passport Issued (dd/mm/yyyy)",
    "Passport Issued Country",
    "Passport Nationality",
]

# Age thresholds (years) used by most GDS / airline pax-type rules.
INFANT_MAX_AGE = 2   # age < 2  -> INF
CHILD_MAX_AGE = 12    # 2 <= age < 12 -> CHD, age >= 12 -> ADT

DEFAULT_NATIONALITY = "Uzbekistan"

TITLES = {
    ("ADT", "Male"): "mr",
    ("ADT", "Female"): "mrs",
    ("CHD", "Male"): "mstr",
    ("CHD", "Female"): "miss",
    ("INF", "Male"): "mstr",
    ("INF", "Female"): "miss",
}

INPUT_HEADER_ALIASES = {
    "full name": "full_name",
    "name": "full_name",
    "date of birth": "dob",
    "dob": "dob",
    "passport №": "passport_no",
    "passport no": "passport_no",
    "passport number": "passport_no",
    "expiry": "expiry",
    "room": "room",
}


def find_header_row(ws, max_scan=10):
    for row in range(1, max_scan + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        normalized = [str(v).strip().lower() for v in values if v is not None]
        if any(v in INPUT_HEADER_ALIASES for v in normalized):
            return row
    raise ValueError("Could not find a header row with a 'Full Name' column in the input file")


def read_input_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws)

    col_map = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=c).value
        if raw is None:
            continue
        key = INPUT_HEADER_ALIASES.get(str(raw).strip().lower())
        if key:
            col_map[key] = c

    required = {"full_name", "dob", "passport_no", "expiry"}
    missing = required - col_map.keys()
    if missing:
        raise ValueError(f"Input file is missing expected column(s): {sorted(missing)}")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        full_name = ws.cell(row=r, column=col_map["full_name"]).value
        if full_name is None or not str(full_name).strip():
            continue
        rows.append({
            "full_name": str(full_name).strip(),
            "dob": ws.cell(row=r, column=col_map["dob"]).value,
            "passport_no": ws.cell(row=r, column=col_map["passport_no"]).value,
            "expiry": ws.cell(row=r, column=col_map["expiry"]).value,
            "room": ws.cell(row=r, column=col_map["room"]).value if "room" in col_map else None,
        })
    return rows


def split_name(full_name):
    """'SURNAME GIVEN NAME(S)' -> (given_name, surname)."""
    parts = full_name.split()
    if len(parts) == 1:
        return parts[0], parts[0]
    last_name = parts[0]
    first_name = " ".join(parts[1:])
    return first_name, last_name


def guess_gender(first_name, last_name):
    # Heuristic: Uzbek/Russian-style female names (and most female surnames)
    # end in "-a". Checking both first and last name catches families whose
    # surname doesn't decline by gender (e.g. "GANIYUT"). Not 100% reliable
    # (rare male names like "Muradulla" also end in "-a") -- review Gender/
    # Title on the output before use.
    return "Female" if first_name.strip().upper().endswith("A") or last_name.strip().upper().endswith("A") else "Male"


def compute_age(dob, ref_date):
    if not isinstance(dob, (dt.date, dt.datetime)):
        raise ValueError(f"Unrecognized date of birth value: {dob!r}")
    if isinstance(dob, dt.datetime):
        dob = dob.date()
    years = ref_date.year - dob.year
    if (ref_date.month, ref_date.day) < (dob.month, dob.day):
        years -= 1
    return years


def pax_type_for_age(age):
    if age < INFANT_MAX_AGE:
        return "INF"
    if age < CHILD_MAX_AGE:
        return "CHD"
    return "ADT"


def parse_travel_date(explicit_date, filename, year):
    if explicit_date:
        return dt.datetime.strptime(explicit_date, "%d.%m.%Y").date()

    m = re.search(r"(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?", filename)
    if not m:
        raise ValueError(
            "Could not find a DD.MM date in the input filename; pass --date DD.MM.YYYY explicitly"
        )
    day, month = int(m.group(1)), int(m.group(2))
    if m.group(3):
        y = int(m.group(3))
        if y < 100:
            y += 2000
    else:
        y = year or dt.date.today().year
    return dt.date(y, month, day)


def build_pax_rows(input_rows, travel_date):
    pax_rows = []
    for i, row in enumerate(input_rows, start=1):
        first_name, last_name = split_name(row["full_name"])
        gender = guess_gender(first_name, last_name)
        age = compute_age(row["dob"], travel_date)
        pax_type = pax_type_for_age(age)
        title = TITLES[(pax_type, gender)]
        pax_rows.append({
            "Sequence": i,
            "Traveling With": None,
            "Pax Type": pax_type,
            "Title": title,
            "First Name": first_name,
            "Last Name": last_name,
            "Gender": gender,
            "DOB (dd/mm/yyyy)": row["dob"],
            "Nationality": DEFAULT_NATIONALITY,
            "Mobile Number": None,
            "Email": None,
            "ID Number": None,
            "Passport Number": row["passport_no"],
            "Passport Expiry (dd/mm/yyyy)": row["expiry"],
            "Passport Issued (dd/mm/yyyy)": None,
            "Passport Issued Country": DEFAULT_NATIONALITY,
            "Passport Nationality": DEFAULT_NATIONALITY,
        })
    return pax_rows


def write_output(pax_rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pax Data"

    for c, header in enumerate(PAX_DATA_HEADERS, start=1):
        ws.cell(row=1, column=c, value=header)

    date_cols = {"DOB (dd/mm/yyyy)", "Passport Expiry (dd/mm/yyyy)", "Passport Issued (dd/mm/yyyy)"}
    for r, row in enumerate(pax_rows, start=2):
        for c, header in enumerate(PAX_DATA_HEADERS, start=1):
            cell = ws.cell(row=r, column=c, value=row[header])
            if header in date_cols and row[header] is not None:
                cell.number_format = "DD/MM/YYYY"

    widths = [10, 14, 10, 8, 20, 20, 10, 16, 14, 14, 20, 12, 16, 20, 20, 18, 18]
    for c, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = width

    # Reference sheet for Pax Type / Title dropdowns, mirroring the source template.
    ref = wb.create_sheet("Reference")
    ref["A1"] = "Pax Type"
    for i, v in enumerate(["ADT", "CHD", "INF"], start=2):
        ref.cell(row=i, column=1, value=v)
    ref["B1"] = "Gender"
    for i, v in enumerate(["Male", "Female"], start=2):
        ref.cell(row=i, column=2, value=v)
    ref["C1"] = "Title"
    for i, v in enumerate(["mr", "mrs", "ms", "mstr", "miss"], start=2):
        ref.cell(row=i, column=3, value=v)
    ref["D1"] = "Nationality"
    ref["D2"] = DEFAULT_NATIONALITY
    ref.sheet_state = "hidden"

    last_row = len(pax_rows) + 1
    if last_row >= 2:
        dv_pax_type = DataValidation(type="list", formula1="=Reference!$A$2:$A$4", allow_blank=True)
        ws.add_data_validation(dv_pax_type)
        dv_pax_type.add(f"C2:C{last_row}")

        dv_gender = DataValidation(type="list", formula1="=Reference!$B$2:$B$3", allow_blank=True)
        ws.add_data_validation(dv_gender)
        dv_gender.add(f"G2:G{last_row}")

        dv_title = DataValidation(type="list", formula1="=Reference!$C$2:$C$6", allow_blank=True)
        ws.add_data_validation(dv_title)
        dv_title.add(f"D2:D{last_row}")

    wb.save(out_path)


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", help="Input xlsx with the simple passenger list")
    parser.add_argument("-o", "--output", help="Output xlsx path (default: <input>_pax_data.xlsx)")
    parser.add_argument("--date", help="Travel date as DD.MM.YYYY (overrides filename/--year)")
    parser.add_argument("--year", type=int, help="Year to use with the DD.MM date found in the filename")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1

    travel_date = parse_travel_date(args.date, input_path.name, args.year)
    input_rows = read_input_rows(input_path)
    pax_rows = build_pax_rows(input_rows, travel_date)

    output_path = Path(args.output) if args.output else input_path.with_name(input_path.stem + "_pax_data.xlsx")
    write_output(pax_rows, output_path)

    print(f"Travel date used for age calculation: {travel_date.isoformat()}")
    print(f"Converted {len(pax_rows)} passengers -> {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
